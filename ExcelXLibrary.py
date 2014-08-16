from openpyxl import load_workbook

__author__ = 'ikabelerma'


class ExcelXLibrary:

    ROBOT_LIBRARY_SCOPE = 'TEST SUITE'

    def open_excel_file(self, path):
        workbook = load_workbook(path)
        return workbook

    def get_cell_value_by_sheet_name(self, workbook, sheet_name, cell_name):
        workbook_sheet = workbook.get_sheet_by_name(sheet_name)
        cell_value = workbook_sheet.cell(cell_name).value
        return cell_value

    def get_cell_value_of_active_sheet(self, workbook, cell_name):
        workbook_sheet = workbook.get_active_sheet()
        cell_value = workbook_sheet.cell(cell_name).value
        return cell_value